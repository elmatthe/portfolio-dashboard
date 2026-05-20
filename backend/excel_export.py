"""Excel export — 5 formatted sheets via openpyxl.

The output is what a non-technical user would expect: dark header rows,
readable currency formatting, green/red conditional fills on gain/loss columns.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from backend import store
from backend.acb import compute as compute_acb
from backend.portfolio import build_portfolio


# ---------- shared styling ----------

HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = Font(color="F9FAFB", bold=True, size=11, name="Calibri")
SECTION_FILL = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
SECTION_FONT = Font(color="F9FAFB", bold=True, size=12)
ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
GAIN_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
BORDER = Border(
    left=Side(border_style="thin", color="E5E7EB"),
    right=Side(border_style="thin", color="E5E7EB"),
    top=Side(border_style="thin", color="E5E7EB"),
    bottom=Side(border_style="thin", color="E5E7EB"),
)


def _money_fmt(currency: str) -> str:
    """Excel money number-format string for the given currency."""
    return f'"$"#,##0.00_-;[Red]-"$"#,##0.00;-' if currency == "CAD" else f'"$"#,##0.00_-;[Red]-"$"#,##0.00;-'


def _pct_fmt() -> str:
    """Excel percent number-format string."""
    return "0.00%;[Red]-0.00%;-"


def _autosize(ws: Worksheet) -> None:
    """Auto-size every column in a worksheet based on the widest cell value."""
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 8
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 38)


def _write_header(ws: Worksheet, row: int, headers: Iterable[str]) -> None:
    """Write a styled header row at the given row index."""
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")


# ---------- main entry ----------

PERIOD_NAMES = {
    "1m": "Last 30 days",
    "3m": "Last 3 months",
    "6m": "Last 6 months",
    "ytd": "Year to date",
    "1y": "Last 12 months",
    "3y": "Last 3 years",
    "all": "All time",
}


def build_xlsx_bytes(period: str | None = None) -> bytes:
    """Build the export file and return its bytes (caller streams to client).

    When `period` is set (1m/3m/6m/ytd/1y/3y/all), the export reflects that
    time window: per-period balances in Sheet 1, transactions filtered to the
    period in Sheet 4, price-history filtered in Sheet 5.
    """
    from backend.portfolio import period_to_start_date, normalize_period

    period_key = normalize_period(period)
    start_date = period_to_start_date(period_key) if period_key != "all" else None
    data = build_portfolio(period=period_key)
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_summary(wb, data, period_key, start_date)
    _sheet_holdings(wb, data, period_key)
    _sheet_capital_gains(wb, data)
    _sheet_transactions(wb, start_date)
    _sheet_price_history(wb, data, start_date)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def suggested_filename(period: str | None = None) -> str:
    """Filename includes the period label so saved exports are self-describing."""
    from backend.portfolio import normalize_period

    key = normalize_period(period)
    suffix = f"_{key}" if key != "all" else ""
    return f"portfolio_export_{date.today().isoformat()}{suffix}.xlsx"


# ---------- sheets ----------

def _sheet_summary(wb: Workbook, data, period_key: str = "all", start_date: date | None = None) -> None:
    """Build the 'summary' worksheet."""
    ws = wb.create_sheet("Portfolio Summary")
    if period_key != "all" and start_date:
        title = (
            f"Portfolio Summary — {PERIOD_NAMES[period_key]} "
            f"({start_date.isoformat()} – {date.today().isoformat()})"
        )
    else:
        title = "Portfolio Summary"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="0F172A")
    ws["A2"] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(color="6B7280", italic=True)
    ws["A3"] = f"USD/CAD: {data.exchange_rate.usd_cad:.4f}"
    ws["A4"] = f"Period: {PERIOD_NAMES.get(period_key, 'All time')}"
    ws["A4"].font = Font(color="6B7280", italic=True)

    _write_header(ws, 5, ["Account", "Cash Deposited CAD", "Cash Deposited USD",
                          "Cash Invested CAD", "Cash Invested USD", "Total Fees CAD",
                          "Total Dividends CAD", "Total Dividends USD",
                          "Total Equity CAD", "Total Equity USD",
                          "Unrealized Gain CAD", "Overall ROI %"])
    r = 6
    for a in data.accounts + [_combined_row(data)]:
        ws.cell(row=r, column=1, value=a.account_type)
        ws.cell(row=r, column=2, value=a.cash_deposited_cad).number_format = _money_fmt("CAD")
        ws.cell(row=r, column=3, value=a.cash_deposited_usd).number_format = _money_fmt("USD")
        ws.cell(row=r, column=4, value=a.cash_invested_cad).number_format = _money_fmt("CAD")
        ws.cell(row=r, column=5, value=a.cash_invested_usd).number_format = _money_fmt("USD")
        ws.cell(row=r, column=6, value=a.total_fees_cad).number_format = _money_fmt("CAD")
        ws.cell(row=r, column=7, value=a.total_dividends_cad).number_format = _money_fmt("CAD")
        ws.cell(row=r, column=8, value=a.total_dividends_usd).number_format = _money_fmt("USD")
        ws.cell(row=r, column=9, value=a.total_equity_cad).number_format = _money_fmt("CAD")
        ws.cell(row=r, column=10, value=a.total_equity_usd).number_format = _money_fmt("USD")
        ws.cell(row=r, column=11, value=a.unrealized_gain_cad).number_format = _money_fmt("CAD")
        roi_cell = ws.cell(row=r, column=12, value=(a.overall_roi_pct or 0) / 100)
        roi_cell.number_format = _pct_fmt()
        if (a.overall_roi_pct or 0) >= 0:
            roi_cell.fill = GAIN_FILL
        else:
            roi_cell.fill = LOSS_FILL
        r += 1
    _autosize(ws)


def _combined_row(data):
    """A row labelled 'Combined' for the bottom of the summary table."""
    c = data.combined.model_copy()
    c.account_type = "Combined"  # type: ignore[assignment]
    return c


def _sheet_holdings(wb: Workbook, data, period_key: str = "all") -> None:
    """Build the 'holdings' worksheet."""
    ws = wb.create_sheet("Holdings")
    headers = ["Ticker", "Security Name", "Account", "Currency", "Shares",
               "ACB / Share", "Total Cost", "Commission", "Current Price",
               "Market Value", "Unrealized Gain", "ROI %",
               "Dividends Received", "Portfolio Weight %"]
    if period_key != "all":
        headers += [f"{period_key.upper()} Return %", f"{period_key.upper()} Dividends"]
    _write_header(ws, 1, headers)
    r = 2
    for h in data.holdings:
        ws.cell(row=r, column=1, value=h.ticker)
        ws.cell(row=r, column=2, value=h.security_name)
        ws.cell(row=r, column=3, value=h.account_type)
        ws.cell(row=r, column=4, value=h.currency)
        ws.cell(row=r, column=5, value=h.total_shares).number_format = "0.####"
        ws.cell(row=r, column=6, value=h.acb_per_share).number_format = _money_fmt(h.currency)
        ws.cell(row=r, column=7, value=h.total_cost).number_format = _money_fmt(h.currency)
        ws.cell(row=r, column=8, value=h.total_commission).number_format = _money_fmt(h.currency)
        ws.cell(row=r, column=9, value=h.current_price).number_format = _money_fmt(h.currency)
        ws.cell(row=r, column=10, value=h.market_value).number_format = _money_fmt(h.currency)
        unrealized_cell = ws.cell(row=r, column=11, value=h.unrealized_gain)
        unrealized_cell.number_format = _money_fmt(h.currency)
        if (h.unrealized_gain or 0) >= 0:
            unrealized_cell.fill = GAIN_FILL
        else:
            unrealized_cell.fill = LOSS_FILL
        roi_cell = ws.cell(row=r, column=12, value=(h.roi_pct or 0) / 100)
        roi_cell.number_format = _pct_fmt()
        roi_cell.fill = GAIN_FILL if (h.roi_pct or 0) >= 0 else LOSS_FILL
        ws.cell(row=r, column=13, value=h.dividends_received).number_format = _money_fmt(h.currency)
        weight_cell = ws.cell(row=r, column=14, value=(h.investment_weight_pct or 0) / 100)
        weight_cell.number_format = _pct_fmt()
        if period_key != "all":
            pr = h.period_return_pct
            pr_cell = ws.cell(row=r, column=15, value=(pr or 0) / 100)
            pr_cell.number_format = _pct_fmt()
            if pr is not None:
                pr_cell.fill = GAIN_FILL if pr >= 0 else LOSS_FILL
            ws.cell(row=r, column=16, value=h.period_dividends_received).number_format = _money_fmt(h.currency)
        r += 1
    _autosize(ws)


def _sheet_capital_gains(wb: Workbook, data) -> None:
    """Build the 'capital gains' worksheet."""
    ws = wb.create_sheet("Capital Gains")
    _write_header(ws, 1, ["Date", "Security", "Account", "Shares Sold", "Sale Price",
                          "ACB / Share", "Gain / Share", "Total Gain", "Commission",
                          "Currency", "Taxable", "Superficial Loss Adj"])
    r = 2
    for g in data.capital_gains.realized_gains:
        ws.cell(row=r, column=1, value=g.transaction_date)
        ws.cell(row=r, column=2, value=g.ticker)
        ws.cell(row=r, column=3, value=g.account_type)
        ws.cell(row=r, column=4, value=g.shares_sold)
        ws.cell(row=r, column=5, value=g.sale_price).number_format = _money_fmt(g.currency)
        ws.cell(row=r, column=6, value=g.acb_per_share).number_format = _money_fmt(g.currency)
        ws.cell(row=r, column=7, value=g.gain_per_share).number_format = _money_fmt(g.currency)
        gain_cell = ws.cell(row=r, column=8, value=g.total_gain)
        gain_cell.number_format = _money_fmt(g.currency)
        gain_cell.fill = GAIN_FILL if g.total_gain >= 0 else LOSS_FILL
        ws.cell(row=r, column=9, value=g.commission).number_format = _money_fmt(g.currency)
        ws.cell(row=r, column=10, value=g.currency)
        ws.cell(row=r, column=11, value="Yes" if g.taxable else "No (Registered)")
        ws.cell(row=r, column=12, value=g.superficial_loss_adjustment).number_format = _money_fmt(g.currency)
        r += 1

    if r > 2:
        r += 1
        ws.cell(row=r, column=1, value="TOTALS").font = Font(bold=True)
        ws.cell(row=r, column=7, value="Total Taxable Gain:").font = Font(bold=True)
        c = ws.cell(row=r, column=8, value=data.capital_gains.total_taxable_gain)
        c.number_format = _money_fmt("CAD")
        c.font = Font(bold=True)
        r += 1
        ws.cell(row=r, column=7, value="Total Non-Taxable:").font = Font(bold=True)
        c = ws.cell(row=r, column=8, value=data.capital_gains.total_non_taxable_gain)
        c.number_format = _money_fmt("CAD")
        c.font = Font(bold=True)
        r += 2
        ws.cell(row=r, column=1, value="Note: 50% inclusion rate applies to taxable capital gains per CRA.")
        ws.cell(row=r, column=1).font = Font(italic=True, color="6B7280")

    _autosize(ws)


def _sheet_transactions(wb: Workbook, start_date: date | None = None) -> None:
    """Build the 'transactions' worksheet."""
    ws = wb.create_sheet("Transaction History")
    _write_header(ws, 1, ["Date", "Broker", "Action", "Symbol", "Resolved Ticker",
                          "Description", "Quantity", "Price", "Commission",
                          "Net Amount", "Currency", "Account #", "Account Type"])
    r = 2
    txs = store.get_all_transactions()
    if start_date is not None:
        txs = [t for t in txs if t.transaction_date >= start_date]
    for t in txs:
        ws.cell(row=r, column=1, value=t.transaction_date)
        ws.cell(row=r, column=2, value=t.broker)
        ws.cell(row=r, column=3, value=t.action)
        ws.cell(row=r, column=4, value=t.raw_symbol)
        ws.cell(row=r, column=5, value=t.resolved_ticker)
        ws.cell(row=r, column=6, value=t.description)
        ws.cell(row=r, column=7, value=t.quantity)
        ws.cell(row=r, column=8, value=t.price).number_format = _money_fmt(t.currency)
        ws.cell(row=r, column=9, value=t.commission).number_format = _money_fmt(t.currency)
        ws.cell(row=r, column=10, value=t.net_amount).number_format = _money_fmt(t.currency)
        ws.cell(row=r, column=11, value=t.currency)
        ws.cell(row=r, column=12, value=t.account_number)
        ws.cell(row=r, column=13, value=t.account_type)
        if r % 2 == 0:
            for col in range(1, 14):
                ws.cell(row=r, column=col).fill = ALT_FILL
        r += 1
    _autosize(ws)


def _sheet_price_history(wb: Workbook, data, start_date: date | None = None) -> None:
    """Build the 'price history' worksheet."""
    ws = wb.create_sheet("Price History")
    col_offset = 1
    for h in data.holdings:
        df = store.get_price_history(h.ticker)
        if df.empty:
            continue
        if start_date is not None:
            import pandas as _pd
            df = df[df.index >= _pd.Timestamp(start_date)]
            if df.empty:
                continue
        col_letter = get_column_letter(col_offset)
        ws.cell(row=1, column=col_offset, value=f"{h.ticker} ({h.security_name or ''})").font = Font(bold=True)
        _write_header(ws, 2, ["Date", "Close", "Weekly Return %"])
        # Re-do headers for this block's actual columns
        ws.cell(row=2, column=col_offset, value="Date").fill = HEADER_FILL
        ws.cell(row=2, column=col_offset).font = HEADER_FONT
        ws.cell(row=2, column=col_offset + 1, value="Close").fill = HEADER_FILL
        ws.cell(row=2, column=col_offset + 1).font = HEADER_FONT
        ws.cell(row=2, column=col_offset + 2, value="Weekly Return %").fill = HEADER_FILL
        ws.cell(row=2, column=col_offset + 2).font = HEADER_FONT

        weekly = df["close"].resample("W").last().dropna()
        returns = weekly.pct_change()
        for i, (idx, close) in enumerate(weekly.items(), start=3):
            ws.cell(row=i, column=col_offset, value=idx.date())
            ws.cell(row=i, column=col_offset + 1, value=float(close)).number_format = _money_fmt(h.currency)
            wr = returns.loc[idx]
            if wr == wr:  # not NaN
                cell = ws.cell(row=i, column=col_offset + 2, value=float(wr))
                cell.number_format = _pct_fmt()
                cell.fill = GAIN_FILL if wr >= 0 else LOSS_FILL

        col_offset += 4  # 3 columns + 1 spacer

    if col_offset == 1:
        ws.cell(row=1, column=1, value="No price history yet — refresh prices and try again.")
        ws.cell(row=1, column=1).font = Font(italic=True, color="6B7280")
    _autosize(ws)
