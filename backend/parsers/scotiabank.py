"""Scotia iTRADE — CSV (semicolon-delimited, DD/MM/YYYY dates) and XLSX.

Quirks:
- Semicolon delimiter, NOT comma.
- Dates in DD/MM/YYYY format — we pass prefer_dmy=True through the
  shared converter so 04/01/2024 reads as Apr 1, not Jan 4.
- Extra `Settlement Date` column → captured.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import ClassVar

import pandas as pd

from backend.models import Transaction
from backend.parsers.base import BaseParser
from backend.parsers.rbc import _rows_to_transactions
from backend.parsers.registry import _register

logger = logging.getLogger(__name__)


@_register
class ScotiabankParser(BaseParser):
    BROKER_NAME: ClassVar[str] = "Scotia iTRADE"
    BROKER_KEY: ClassVar[str] = "scotiabank"
    SUPPORTED_FORMATS: ClassVar[list[str]] = ["csv", "xlsx"]

    @classmethod
    def detect(cls, file_path: str | Path, content_sample: str) -> float:
        ext = Path(file_path).suffix.lower()
        s = content_sample
        if ext == ".csv":
            first_line = s.splitlines()[0] if s else ""
            if ";" in first_line and "Order Type" in first_line and "Book Value CAD" in first_line:
                return 0.9
            if "itrade" in s.lower():
                return 0.95
        if ext in (".xlsx", ".xlsm"):
            try:
                from openpyxl import load_workbook

                wb = load_workbook(file_path, read_only=True, data_only=True)
                if any("trade history" in (s or "").lower() for s in wb.sheetnames):
                    return 0.9
            except Exception:
                return 0.0
        return 0.0

    def parse(self, file_path: str | Path) -> list[Transaction]:
        path = Path(file_path)
        ext = path.suffix.lower()
        if ext == ".csv":
            df = pd.read_csv(path, sep=";", encoding_errors="replace")
            return _rows_to_transactions(
                df, "scotiabank", broker="scotiabank",
                default_currency="CAD", prefer_dmy=True,
            )
        if ext in (".xlsx", ".xlsm"):
            return _parse_scotia_xlsx(path)
        return []


def _parse_scotia_xlsx(path: Path) -> list[Transaction]:
    """Scotia XLSX is multi-sheet (Trade History + Account Summary). Read
    ONLY the Trade History sheet. Headers are on row 1, data starts row 2.

    The `Order` column carries the action label (Buy/Sell/etc.) — our
    `_COL_ALIASES["action"]` already includes "order" via "order type", but
    the bare "Order" string isn't there; we map it manually below.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if "trade history" in s.lower()), wb.sheetnames[0])
    ws = wb[sheet_name]

    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    # Rename "Order" → "Order Type" so the existing alias picks it up.
    headers = ["Order Type" if h.strip().lower() == "order" else h for h in headers]
    rows = []
    for r in range(2, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if first is None:
            continue
        rec = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
        rows.append(rec)
    df = pd.DataFrame(rows)
    # Scotia XLSX gives Python datetime objects so prefer_dmy doesn't matter
    # for the date column — but harmless to keep on for textual fallbacks.
    return _rows_to_transactions(
        df, "scotiabank", broker="scotiabank",
        default_currency="CAD", prefer_dmy=False,
    )
