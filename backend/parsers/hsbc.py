"""HSBC InvestDirect — XLSX.

Quirks:
- Each row carries an ISIN (international security identifier) — captured.
- `Local Currency` is the trade currency; `Settlement CCY` is the
  settlement currency, usually CAD. We use Local Currency for the
  Transaction.currency field.
- The file already includes `FX to CAD` and `CAD Value` columns — both
  are taken at face value (priority-1 source per §3.2).
- 33 transactions; single `Transactions` sheet.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import ClassVar

import pandas as pd

from backend.fx import get_fx_service
from backend.models import Transaction
from backend.parsers._common import (
    compute_hash,
    guess_account_type,
    guess_currency,
    guess_exchange,
    normalize_action,
    parse_date,
    safe_float,
    safe_str,
)
from backend.parsers.base import BaseParser
from backend.parsers.registry import _register

logger = logging.getLogger(__name__)


@_register
class HSBCParser(BaseParser):
    BROKER_NAME: ClassVar[str] = "HSBC InvestDirect"
    BROKER_KEY: ClassVar[str] = "hsbc"
    SUPPORTED_FORMATS: ClassVar[list[str]] = ["xlsx"]

    @classmethod
    def detect(cls, file_path: str | Path, content_sample: str) -> float:
        ext = Path(file_path).suffix.lower()
        if ext not in (".xlsx", ".xlsm"):
            return 0.0
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(ws.cell(1, c).value or "").lower() for c in range(1, ws.max_column + 1)]
            joined = " ".join(headers)
            if "isin" in joined and "local currency" in joined and "settlement ccy" in joined:
                return 0.95
            if "investdirect" in joined or "hsbc" in joined:
                return 0.9
        except Exception:
            return 0.0
        return 0.0

    def parse(self, file_path: str | Path) -> list[Transaction]:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)
        ws = wb["Transactions"] if "Transactions" in wb.sheetnames else wb.active

        headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        rows: list[dict] = []
        for r in range(2, ws.max_row + 1):
            first = ws.cell(r, 1).value
            if first is None:
                continue
            rec = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
            rows.append(rec)

        fx = get_fx_service()
        out: list[Transaction] = []
        for rec in rows:
            trade_date = parse_date(rec.get("Value Date"))
            if trade_date is None:
                continue
            isin = safe_str(rec.get("ISIN"))
            symbol = safe_str(rec.get("Ticker"))
            description = safe_str(rec.get("Name")) or ""
            currency = guess_currency(rec.get("Local Currency"), default="USD")
            quantity = safe_float(rec.get("Quantity"))
            price = safe_float(rec.get("Trade Price"))
            local_amount = safe_float(rec.get("Local Amount"))
            net_settlement = safe_float(rec.get("Net Settlement"))
            file_fx = safe_float(rec.get("FX to CAD")) or None
            cad_value = safe_float(rec.get("CAD Value")) or None
            account_label = safe_str(rec.get("Account")) or "Non-Registered"
            account_type = guess_account_type(account_label, default="Non-Registered")

            raw_action = safe_str(rec.get("Transaction"))
            action = normalize_action(raw_action, quantity=quantity, net=local_amount)
            # Buys cost cash — re-sign local_amount and cad_value if needed.
            if action == "BUY" and local_amount > 0:
                local_amount = -local_amount
                if cad_value and cad_value > 0:
                    cad_value = -cad_value

            fx_rate = file_fx if file_fx else fx.rate_to_cad(currency, trade_date)
            # Stash the file's per-row FX as an "in-file" override for future imports.
            if file_fx:
                fx.register_in_file_rate(currency, trade_date, file_fx)
            if cad_value is None:
                cad_value = round(local_amount * fx_rate, 2)

            account_number = f"hsbc-{account_type}"
            h = compute_hash(
                transaction_date=trade_date,
                action=action,
                raw_symbol=symbol,
                quantity=quantity,
                net_amount=local_amount,
                account_number=account_number,
            )
            out.append(
                Transaction(
                    hash=h,
                    broker="hsbc",
                    transaction_date=trade_date,
                    action=action,
                    raw_symbol=symbol,
                    description=description,
                    quantity=quantity,
                    price=price,
                    gross_amount=abs(local_amount),
                    commission=0.0,  # HSBC bundles commission into the net
                    net_amount=local_amount,
                    currency=currency,
                    account_number=account_number,
                    account_type=account_type,
                    fx_rate_to_cad=fx_rate,
                    net_cad=cad_value,
                    isin=isin,
                    exchange=guess_exchange(symbol, currency),
                )
            )
        return out
