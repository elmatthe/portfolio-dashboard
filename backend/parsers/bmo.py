"""BMO InvestorLine — CSV and XLSX.

CSV quirks:
- Fully quoted CSV.
- 3-line metadata header + blank + column row.
- `Account Type` is the FIRST column (not last); column order inverted vs
  the other Canadian brokers.

XLSX (Step 5) has merged title cells in rows 1-3 and a summary formula row
at the bottom — handled separately by the openpyxl-based xlsx parser.
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import ClassVar

import pandas as pd

from backend.models import Transaction
from backend.parsers.base import BaseParser
from backend.parsers.rbc import _find_csv_header_row, _rows_to_transactions
from backend.parsers.registry import _register

logger = logging.getLogger(__name__)


@_register
class BMOParser(BaseParser):
    BROKER_NAME: ClassVar[str] = "BMO InvestorLine"
    BROKER_KEY: ClassVar[str] = "bmo"
    SUPPORTED_FORMATS: ClassVar[list[str]] = ["csv", "xlsx"]

    @classmethod
    def detect(cls, file_path: str | Path, content_sample: str) -> float:
        ext = Path(file_path).suffix.lower()
        s = content_sample.lower()
        if ext == ".csv":
            if "bmo investorline" in s:
                return 0.95
            if "trans. date" in s and "no. of shares" in s:
                return 0.8
        if ext in (".xlsx", ".xlsm"):
            # XLSX detect is finalised in Step 5 (openpyxl peek at the merged
            # title row in row 1).
            try:
                from openpyxl import load_workbook

                wb = load_workbook(file_path, read_only=True, data_only=True)
                first_cell = (wb.active.cell(1, 1).value or "")
                return 0.9 if "bmo" in str(first_cell).lower() else 0.0
            except Exception:
                return 0.0
        return 0.0

    def parse(self, file_path: str | Path) -> list[Transaction]:
        path = Path(file_path)
        ext = path.suffix.lower()
        if ext == ".csv":
            return _parse_bmo_csv(path)
        if ext in (".xlsx", ".xlsm"):
            return _parse_bmo_xlsx(path)
        return []


def _parse_bmo_csv(path: Path) -> list[Transaction]:
    text = path.read_text(encoding="utf-8", errors="replace")
    account_number = "bmo"
    for line in text.splitlines()[:4]:
        m = re.search(r'"Account",\s*"([\w-]+)"', line)
        if m:
            account_number = m.group(1)
            break
    header_idx = _find_csv_header_row(text, must_contain=("Trans. Date", "Net Proceeds"))
    if header_idx is None:
        logger.warning("BMO parser: no header row found in %s", path)
        return []
    df = pd.read_csv(path, skiprows=header_idx)
    return _rows_to_transactions(df, account_number, broker="bmo", default_currency="CAD")


def _parse_bmo_xlsx(path: Path) -> list[Transaction]:
    """BMO XLSX has merged title cells in rows 1-3 and a summary row at the
    bottom. Headers live on row 5; data starts at row 6.

    We use openpyxl directly so merged cells don't confuse pandas, then hand
    the resulting DataFrame to the shared row converter.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    if "Transactions" in wb.sheetnames:
        ws = wb["Transactions"]
    else:
        ws = wb.active

    # Locate the column-header row by scanning for one whose first cell is
    # "Trade Date" / "Trans. Date" / "Date".
    header_row_idx = None
    for r in range(1, min(ws.max_row, 10) + 1):
        first = ws.cell(r, 1).value
        if isinstance(first, str) and first.strip().lower() in {"trade date", "trans. date", "transaction date", "date"}:
            header_row_idx = r
            break
    if header_row_idx is None:
        logger.warning("BMO XLSX parser: header row not found in %s", path)
        return []
    headers = [str(ws.cell(header_row_idx, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if first is None or (isinstance(first, str) and first.strip().lower() in {"total", "summary", ""}):
            continue
        rec = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
        rows.append(rec)

    # Pull account number from the merged metadata cell on row 3 if present.
    account_number = "bmo"
    for r in range(1, header_row_idx):
        for c in range(1, ws.max_column + 1):
            cell_val = ws.cell(r, c).value
            if isinstance(cell_val, str):
                m = re.search(r"Account[:\s]+([\w-]+)", cell_val)
                if m:
                    account_number = m.group(1)
                    break
        if account_number != "bmo":
            break

    df = pd.DataFrame(rows)
    return _rows_to_transactions(df, account_number, broker="bmo", default_currency="CAD")
